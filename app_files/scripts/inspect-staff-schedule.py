#!/usr/bin/env python3
"""Inspect the staff A/B schedule workbook shape used by exports."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: python scripts/inspect-staff-schedule.py '/path/to/2025 STAFF AB SCHEDULE.xlsx'")
        return 2

    workbook = openpyxl.load_workbook(Path(sys.argv[1]), data_only=True)
    summary = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
        summary.append(
            {
                "sheet": sheet_name,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "headers": headers,
                "sampleRows": [
                    [sheet.cell(row, column).value for column in range(1, sheet.max_column + 1)]
                    for row in range(2, min(sheet.max_row, 6) + 1)
                ],
            }
        )

    print(json.dumps(summary, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
